from scipy.optimize import minimize, LinearConstraint, NonlinearConstraint, Bounds
from scipy.integrate import solve_ivp
import numpy as np
import pandas as pd
import itertools
import math
import os
import re
import openpyxl

STANDARD_RATE_WT = .377/.1/580
#STANDARD_RATE_VOL = 
WATER_PPG = 8.3454

#Not yet implemented - designed to be used with custom temperature curves or cost functions to allow math equations to use python code
def test_func(func, allowed_vars):
        """
        Test a given cost or temperature function with a set of allowed variables to prevent arbitrary code execution. Tokens are limited to those absolutely necessary for math functions and math/numpy library calls.

        Args:
            func: The function to test (as a string of python code)
        """
        allowed_modules = ["np", "math"]
        allowed_functions = ["exp", "log", "sin", "cos", "tan", "asin", "acos", "atan"]
        allowed_tokens = [tokenize.ENCODING, tokenize.NAME, tokenize.NUMBER, tokenize.LPAR, tokenize.RPAR, tokenize.COMMA,
                          tokenize.PLUS, tokenize.MINUS, tokenize.STAR, tokenize.SLASH, tokenize.LESS,
                          tokenize.GREATER, tokenize.DOT, tokenize.EQEQUAL, tokenize.NOTEQUAL, tokenize.LESSEQUAL,
                          tokenize.GREATEREQUAL, tokenize.DOUBLESTAR, tokenize.DOUBLESLASH, tokenize.AT, tokenize.COMMENT,
                          tokenize.NEWLINE, tokenize.ENDMARKER]
        tokens = list(tokenize.tokenize(BytesIO(func.encode("UTF-8")).readline))
        for i, token in enumerate(tokens):
                if token.exact_type not in allowed_tokens:
                        raise ValueError(f"Forbidden token {token['exact_type']} in function {k}")
                if token.exact_type != tokenize.ENCODING:
                        if token.exact_type == tokenize.NAME:
                                if i + 1 < len(tokens) and tokens[i+1].exact_type == tokenize.LPAR:
                                        if token.string not in allowed_functions:
                                            raise ValueError(f"Forbidden function call to {token.string} in function {k}")
                                elif i + 1 < len(tokens) and tokens[i+1].exact_type == tokenize.DOT:
                                        if token.string not in allowed_modules:
                                            raise ValueError(f"Forbidden module {token.string} in function {k}")
                                elif token.string not in allowed_vars:
                                        raise ValueError(f"Forbidden variable {token.string} in function {k}")
                                
# also unimplemented, designed to be used with the function abive

def parse_functions(config):
        ret = dict()
        for k in ["hansen_cost", "evap_cost", "voc_cost"]:
                func = config.get(k)
                if func is None:
                        continue
                test_func(func, ["evap_curve", "control_evap_curve", "RED", "control_RED", "VOC", "max_voc", "control_VOC"])
                ret[k] = func
        for curve in config.get("temp_curves"):
                func = curve["func"]
                allowed_vars = curve["vars"].keys()
                test_func(func, allowed_vars)

def get_RER(temperature, blend):
        """
        Get temperature-adjusted relative evaporation rate (RER, nBuAc=100) using Antoine's Equation to determine vapor pressure enhancement and applying it to the 25 C RER.

        Args:
                temperature: A numpy array of temperatures at all time points of interest
                blend: A list of pandas DataFrame rows containing information about the solvents in the blend

        Returns:
                An array of RERs for the provided solvents, in the same order as `blend`
                
        """
        return np.array([10 ** (s["AA"]-s["AB"]/(temperature+s["AC"]))/s["VP@25"]*s["RER"] for s in blend])

def expand_multi_comp(conc, blend, all_solvents):
        """
        Expand multi-component solvents into their individual components, adjusting both `conc` and `blend`
        """
        conc = list(itertools.chain.from_iterable([[c*float(a) for a in s["Amounts"].split(",")] if s["Multi-Component"] else [c] for c, s in zip(conc, blend)]))
        blend = list(itertools.chain.from_iterable([[all_solvents.loc[c.strip(" "),:] for c in s["Components"].split(",")] if s["Multi-Component"] else [s] for s in blend]))
        return conc, blend
        
def get_evap_curve(conc, blend, target, temp_curve, t_span, all_solvents, output_weight, **kwargs):
        """
        Use `scipy.integrate.solve_ivp` to get an evaporation curve. Additional keyword args are passed to `solve_ivp`.

        Args:
                conc: Initial concentration of solvents in blend (volume fraction)
                blend: A list of pandas DataFrame rows containing information about the solvents in the blend
                target: A dictionary containing information about the Hansen Solubility Parameters of the targeted resin (dD, dP, dH, and R0)
                temp_curve: A function which returns the temperature at a given time (in minutes)
                t_span: 2-tuple of start and end times (in minutes, start time is almost always 0)
                all_solvents: Solvent database
                output_weight: Boolean, whether to convert all of the volume fractions into weight fractions at the end.

        Returns:
                tuple containing

                - t: Numpy array of time points where concentrations were evaluated
                - total_profile: Numpy array of total amount of solvent remaining at each time point
                - partial_profiles: 2D Numpy array of the fraction of total remaining blend for each solvent at each time point
                - RED: Numpy array of RED at each time point
        
        """
        conc, blend = expand_multi_comp(conc, blend, all_solvents)
        def dcdt(t, y):
                RER = get_RER(temp_curve(t), blend)
                standard_rate = .377/.1/580
                current_profile = np.clip(y, a_min=0, a_max=np.inf)/sum(np.clip(y, a_min=0, a_max=np.inf))
                return -1 * np.minimum(current_profile * RER * (1-target["solids"]/100)**2 * standard_rate, y/0.01)
        
        try:
                ret = solve_ivp(dcdt, t_span, conc, max_step=0.01, **kwargs)
        except ValueError:
                print(conc, blend, target, temp_curve, t_span)
                raise ValueError
        t = ret["t"][:-1]
        c = np.transpose(ret["y"])
        RED = RED_mix(np.array([r/sum(r) for r in c]), blend, target)[:-1]
        if output_weight:
              c = c * np.tile(np.array([s["Density"] for s in blend]), (len(c), 1))
              c = c / sum(c[0])
        total_profile = np.array([sum(r) for r in c])
        partial_profiles = np.array([r/sum(r) for r in c]).T
        partial_profiles = np.clip(partial_profiles, 0, 1)
        partial_profiles[np.where((abs(np.diff(partial_profiles)) < 1e-10) & (partial_profiles[:,:-1] < 0.5))] = 0
        partial_profiles[np.where((abs(np.diff(partial_profiles)) < 1e-10) & (partial_profiles[:,:-1] > 0.5))] = 1
        partial_profiles = partial_profiles[:,:-1]
        return t, total_profile[:-1], partial_profiles, RED

def get_evap_estimate(conc, blend, target, temp_curve, t_span, n=3, **kwargs):
        """
        Non-differential equation method of estimating evaporation curves.
        Uses an iterative method by guessing partial profiles from total profile.
        Starts with total_profile = 1 and runs for `n` iterations.
        Same Arguments/Return value as get_evap_curve, except we don't care as much about accuracy so conc can be weight or volume.
        """
        t = np.arange(*t_span, (t_span[1]-t_span[0])/100)
        standard_rate = .377/.1/580
        R = get_RER(temp_curve(t), blend) * (1-target["solids"]/100)**2 * standard_rate
        time_profile = np.array([c*np.exp(-Ri*t) for Ri, c in zip(R, conc)])
        total_profile = np.sum(time_profile, axis=0)
        for _ in range(1, n):
                time_profile = np.array([c*np.exp(-Ri*np.cumsum(1/total_profile * (t_span[1]-t_span[0])/100)) for Ri, c in zip(R, conc)])
                total_profile = np.maximum(np.sum(time_profile, axis=0), 1e-20)
        try:
                final_index = np.where(total_profile==1e-20)[0][0]
        except Exception:
                final_index = len(total_profile)

        partial_profiles = np.array([tp[:final_index]/total_profile[:final_index] for tp in time_profile])
        RED = RED_mix(partial_profiles.T, blend, target)
        return t[:final_index], total_profile[:final_index], partial_profiles, RED

def RED_mix(conc, blend, target):
        """
        Calculates RED from a solvent blend to a targeted resin.

        Args:
                conc: Initial concentration of solvents in blend (weight %)
                blend: A list of pandas DataFrame rows containing information about the solvents in the blend
                target: A dictionary containing information about the Hansen Solubility Parameters of the targeted resin (dD, dP, dH, and R0)
                
        Returns:
               Relative Energy Difference (RED) between solvent blend and target resin
        """
        blend_params = dict()
        for p in ["dD", "dP", "dH"]:
                solvent_params = np.array([s[p.replace('d','δ')] for s in blend]).T
                blend_params[p] = conc @ solvent_params
        return ((4 * (blend_params["dD"] - target["dD"])**2 + (blend_params["dP"] - target["dP"])**2 + (blend_params["dH"] - target["dH"])**2)**.5)/target["R0"]

def total_cost(all_solvents, conc, blend, control_params, target, temp_curve, num_exempt, max_voc, estimate_evap = False, debug = False):
        """
        Calculate the total cost of a blend relative to a given control.

        Args:
                all_solvents: DataFrame with solvent information for all possible solvents
                conc: Initial concentration of solvents in blend (weight %)
                blend: A list of pandas DataFrame rows containing information about the solvents in the blend
                control_params: Tuple returned from get_evap_curve or get_evap_estimate for a control blend (for evaporation curve comparison/cost calculation)
                target: A dictionary containing information about the Hansen Solubility Parameters of the targeted resin (dD, dP, dH, and R0)
                temp_curve: A function which returns the temperature at a given time (in minutes)
                max_voc: Maximum amount of VOC permitted in the blend (same units as conc)
                estimate_evap: Whether to estimate the evaporation curve using `get_evap_estimate`. If `false`, `get_evap_curve` will be used
                debug: Enables debugging print statements
                
        Returns:
                Total cost of this alternative blend vs the control
                
        """
        conc, blend = expand_multi_comp(conc, blend, all_solvents)
                        
        if estimate_evap:
                t, total_profile, partial_profiles, RED = get_evap_estimate(conc, blend, target, temp_curve, [0, 10])
        else:
                t, total_profile, partial_profiles, RED = get_evap_curve(conc, blend, target, temp_curve, [0, 10])
        total_profile = np.array(total_profile)
        
        if max(RED) > 1:
                hansen_cost = (max(RED)-0.99) * 10000 # Heavily dissuade this
        else:
                weights = np.clip(total_profile, a_min=0.01, a_max=None) * 100
                hansen_cost = np.dot(np.transpose(1/(1.01 - RED)), weights) / sum(weights) # Typical Cost: 3-5
        clipped_total_profile = np.minimum(np.array(list(total_profile)+[1e-20]*(len(control_params[1])-len(total_profile))), 0.05)
        clipped_control_profile = np.minimum(np.array(list(control_params[1])+[1e-20]*(len(total_profile)-len(control_params[1]))), 0.05)
        #Note: This is a decent cost function for mid-to-late stage evaporation while pretty much completely ignoring initial evaporation
        #Might want to tweak this to have a separate initial component wihtout completely overshadowing this.
        evap_cost = (sum(20*(clipped_total_profile - clipped_control_profile))**2 + sum((20*(clipped_total_profile - clipped_control_profile))**2))*7 #Typical cost: 7-10
        
        wt_voc = sum([c if not comp["Exempt"] else 0 for c, comp in zip(conc, blend)])
        ratio = wt_voc/max_voc
        if ratio < 0.9:
                voc_cost = 0
        else:
                voc_cost = (ratio - 0.9) ** 2 * 100 # Typical cost: 0-1
        return hansen_cost + evap_cost + voc_cost

def get_alternative_blends(all_solvents, control_blend, min_comp, replace_by, target, temp_curve, exempt_range, ne_range, voc_limit, control_density, whitelist = None, blacklist = None):
        """
        Finds the best alternative blends (ranked by total cost) for a given control

        Args:
                all_solvents: DataFrame with solvent information for all possible solvents
                control_blend: A list of pandas DataFrame rows containing information about the solvents in the control blend
                min_comp: DataFrame of minimum composition (included in resins, driers, tints, etc.) with columns for solvents and weight and/or volume fractions
                replace_by: either "Weight Fraction" or "Volume Fraction", directing the function on which units to do optimization in.
                target: A dictionary containing information about the Hansen Solubility Parameters of the targeted resin (dD, dP, dH, and R0)
                temp_curve: A function which returns the temperature at a given time (in minutes)
                exempt_range: 2-tuple of (min, max) for number of exempt solvents to use in alternative blends
                ne_range: 2-tuple of (min, max) for number of non-exempt solvents to use in alternative blends

        Returns:
                List of dictionaries containing information about alternative blends, sorted by cost (low to high).
                Each dictionary contains entries for `conc` and `blend` (as defined in other functions), as well as the number of exempt solvents `num_exempt`,
                total cost `cost`, and `order`, which is just `blend` sorted by highest to lowest `conc`.
        """
        
        min_exempt, max_exempt = exempt_range
        min_ne, max_ne = ne_range
        max_voc, voc_limit_type = voc_limit

        
        if whitelist is None:
                whitelist = all_solvents.index[~all_solvents.index.isin(blacklist)]
                
        exempt_solvents = [s for _, s in all_solvents.loc[all_solvents.Exempt & all_solvents.index.isin(whitelist)].iterrows()]
        ne_solvents = [s for _, s in all_solvents.loc[~all_solvents.Exempt & all_solvents.index.isin(whitelist)].iterrows()]
        control_estimate = get_evap_estimate(control_blend[replace_by], [all_solvents.loc[name,:] for name in control_blend["Name"]], target, temp_curve, [0, 10])

        solvent_density = control_blend["Volume Fraction"] * all_solvents.loc[control_blend["Name"],"Density"]
        solids_vpg = 1 - control_density * (1-target["solids"]/100) / solvent_density
        solvent_vpg = 1 - solids_vpg
        solvent_wpg = (1-target["solids"]/100) * control_density

        #Step 1: Find exempt blends that work decently well
        
        exempt_blends = list(itertools.chain.from_iterable([list(itertools.combinations(exempt_solvents, n)) for n in range(min_exempt, max_exempt + 1)]))
        exempt_results = []
        for eb in exempt_blends:
                if len(eb) == 1:
                        exempt_results.append({"blend":eb, "conc": [1], "cost":total_cost(all_solvents, [1], eb, control_estimate, target, temp_curve, 1, max_voc, True, True)})
                else:
                        sum_constraint = LinearConstraint(np.ones(len(eb)-1), ub=0.9+1e-10)
                        res = minimize(lambda x: total_cost(all_solvents, list(x)+[1-sum(x)], eb, control_estimate, target, temp_curve, len(eb), max_voc, True),
                                       x0 = np.ones(len(eb)-1)/len(eb), bounds = [(0.1,1)]*(len(eb)-1), constraints=sum_constraint)
                        if res.success:
                                exempt_results.append({"blend":eb, 
                                        "conc": list(res.x) + [1-sum(res.x)],
                                        "cost":total_cost(all_solvents, list(res.x) + [1-sum(res.x)], eb, control_estimate, target, temp_curve, len(eb), max_voc, True, True),
                                        })
        good_exempt = sorted(exempt_results, key=lambda x: x["cost"])

        #Step 2 - Use good exempt blends to create ful solvent blend
        
        ne_blends = list(itertools.chain.from_iterable([list(itertools.combinations(ne_solvents, n)) for n in range(min_ne, max_ne + 1)]))
        mc_tot = sum(min_comp[replace_by])
        mc_exempt = sum([r[replace_by] if all_solvents["Exempt"][r["Name"]] else 0 for _, r in min_comp.iterrows()])
        if voc_limit_type in ["vol %", "wt %"]:
                mc_voc = mc_tot - mc_exempt
        """
        Code for lbs/gal limit information - not well tested
        else:
                if voc_limit_type == "g/L":
                        max_voc *= 0.0083 # Convert to lbs/gal
                if replace_by == "Weight Fraction":
                        mc_voc = (mc_tot - mc_exempt) * solvent_wpg #Minimum composition VOC weight (in lbs) per gallon of total formulation
                else:
                        mc_voc = sum([r[replace_by]*all_solvents["Density"][r["Name"]] if all_solvents["Exempt"][r["Name"]] else 0 for _, r in min_comp.iterrows()]) * solvent_wpg
        """
        results = []
        for ge in good_exempt[:min(max(len(good_exempt)//2,3), len(good_exempt))]:
                e = ge["blend"]
                for n in ne_blends:
                        full_blend = list(e) + list(n)
                        entire_blend = full_blend + [all_solvents.loc[n,:] for n in min_comp["Name"]]
                        sum_constraint = LinearConstraint(np.ones(len(full_blend)-1), ub=(1-mc_tot-0.01)+1e-10)
                        if voc_limit_type in ["vol %", "wt %"]:
                                exempt_constraint = LinearConstraint(np.array([int(all_solvents["Exempt"][s.name]) for s in full_blend[:-1]]), lb = 1-max_voc-mc_exempt)
                        """
                        Code for lbs/gal limit information - not well tested
                        elif replace_by == "Weight Fraction":
                                exempt_constraint = LinearConstraint(np.array([int(not all_solvents["Exempt"][s.name]) for s in full_blend[:-1]]) * solvent_wpg, ub = max_voc * solids_vpg - mc_voc)
                        else:
                                exempt_constraint = LinearConstraint(np.array([int(not all_solvents["Exempt"][s.name]) * all_solvents["Density"][s.name] for s in full_blend[:-1]]), ub = max_voc * solids_vpg - mc_voc)
                        """     
                        res = minimize(lambda x: total_cost(all_solvents, list(x)+[1-mc_tot-sum(x)]+list(min_comp[replace_by]), entire_blend , control_estimate, target, temp_curve,
                                       len(e), max_voc, True), x0=[x*(1-mc_exempt-max_voc) for x in ge["conc"]]+[(max_voc-mc_voc)/len(n)]*(len(n)-1),
                                       bounds = [(0.01, 1-mc_tot)]*len(e) + [(0.01, max_voc)] * (len(n)-1), constraints = [sum_constraint, exempt_constraint])
                        if res.success:
                                conc = list(res.x) + [1-sum(res.x)-mc_tot]
                                results.append({"blend":full_blend, 
                                        "conc": conc,
                                        "num_exempt": len(e),
                                        "order": [full_blend[conc.index(x)] for x in sorted(conc, key=lambda x:-x)],
                                        "cost":total_cost(all_solvents, list(res.x) + [1-mc_tot-sum(res.x)] + list(min_comp[replace_by]), entire_blend, control_estimate, target, temp_curve, len(e), max_voc, True, True),
                                        })
        sorted_results = sorted(results, key=lambda x:x["cost"])
        return list(filter(lambda x: x["cost"] <= sorted_results[0]["cost"] * 2, sorted_results))
                



def group_similar_results(results, num_solvents):
        """
        Creates a nested dictionary of results in the same format as is displayed in the "Alternative Blends" tab of the reformulation window

        Args:
                results: Results of alternative blend search from `get_alternative_blends`
                num_solvents: 2-tuple of (min, max) for number of solvents used in alternative blends

        Returns:
               Nested dictionary with higher-conc solvents as keys and lower-conc solvents or results as values
        """
        grouped_results = dict()
        min_n, max_n = num_solvents
        for i in range(min_n, max_n+1):
                for r in filter(lambda x: len(x["conc"]) == i, results):
                        current_level = grouped_results
                        fewer_better = False
                        for solvent in r["order"][:-1]:
                                if current_level.get("result") is not None and current_level["result"]["cost"] < r["cost"]*1.02:
                                        fewer_better = True
                                        break
                                if solvent.name not in current_level:
                                        current_level[solvent.name]=dict()
                                current_level = current_level[solvent.name]
                        if current_level.get("result") is not None and current_level["result"]["cost"] < r["cost"]*1.02:
                                fewer_better = True
                        if not fewer_better:
                                current_level[r["order"][-1].name] = {"result": r}
        return grouped_results



def write_to_excel(fname, blend, t, total_profile, partial_profiles, RED, temperatures, target_params, temp_profile, temp_params, caption=""):
    """
    Creates a new Excel sheet with "summary", "target", and full data tabs based on the outputs of `get_evap_estimate` or `get_evap_curve` for a given `blend`

    Args:
        fname: The filename to write results to
        blend: A list of pandas DataFrame rows containing information about the solvents in the blend
        t: Numpy array of time points where concentrations were evaluated
        total_profile: Numpy array of total amount of solvent remaining at each time point
        partial_profiles: 2D Numpy array of the fraction of total remaining blend for each solvent at each time point
        RED: Numpy array of RED at each time point
        temperatures: A numpy array of temperatures at all time points of interest
        target_params: A dictionary containing information about the Hansen Solubility Parameters of the targeted resin (dD, dP, dH, R0, and solids)
        temp_profile: The selected temperature profile
        temp_params: A dictionary containing the parameters of the selected temperature profile
        caption: A string to be used in the "All Data" sheet name
    """
    if fname[-5:] != ".xlsx":
        fname = fname + ".xlsx"
    # Ensure at least one sheet is visible
    wb = openpyxl.Workbook()
    wb.create_sheet("Summary")
    wb.create_sheet("Target")
    wb.create_sheet("All Data")
    wb.save(fname)

    with pd.ExcelWriter(fname) as writer:
        full_data = {"Time (min)": t}
        full_data.update({name: pp for pp, name in zip(partial_profiles, blend)})
        full_data["Total"] = total_profile
        full_data["RED"] = RED.flatten()
        full_data["Temp (C)"] = temperatures
        full_data_df = pd.DataFrame.from_dict(full_data)
        
        # Modified summary creation
        summary_data = []
        for x in [1, 0.75, 0.5, 0.25, 0.1]:
            row = full_data_df.loc[full_data_df['Total'] <= x].iloc[0] if any(full_data_df['Total'] <= x) else full_data_df.iloc[-1]
            summary_data.append(row.drop(["Total"] + list(blend)))
        summary_df = pd.concat(summary_data, axis=1).T
        summary_df.index = ["100% Solvent Remaining", "75% Solvent Remaining", "50% Solvent Remaining", "25% Solvent Remaining", "10% Solvent Remaining"]
        
        summary_df.to_excel(writer, sheet_name="Summary")
        
        # Add temperature profile information to target_params
        target_params['Temperature Profile'] = temp_profile
        for param, value in temp_params.items():
            target_params[f'Temp Profile - {param}'] = value
        
        target_df = pd.DataFrame.from_dict(target_params, orient='index', columns=['Value'])
        target_df.to_excel(writer, sheet_name="Target")
        
        full_data_df.to_excel(writer, sheet_name=f"All Data")

    print(f"Excel file saved successfully: {fname}")

